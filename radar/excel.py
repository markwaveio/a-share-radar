"""Excel 输出。严格对齐需求方给的 Excel 模板的表头结构。

需求 1 的表头是两行合并表头：第一行是周期名（跨 3 列合并），第二行是 K值/D值/J值。
用 openpyxl 手写合并单元格，不用 pandas.to_excel —— 后者做不出双层表头。
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import TIMEFRAME_COLUMNS

log = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_SUB_FILL = PatternFill("solid", fgColor="DCE6F1")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_SUB_FONT = Font(bold=True, size=10)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header(ws, row: int, ncols: int, fill, font) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, _CENTER, _BORDER


def write_kdj_sheet(ws, rows: Sequence[dict]) -> None:
    """需求 1：KDJ 全周期表。两行表头 + 每只股票一行、30 个数值列。"""
    ws.cell(row=1, column=1, value="个股代码")
    ws.cell(row=1, column=2, value="个股名称")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    col = 3
    for label, _source, _offset in TIMEFRAME_COLUMNS:
        ws.cell(row=1, column=col, value=label)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        for i, sub in enumerate(("K值", "D值", "J值")):
            ws.cell(row=2, column=col + i, value=sub)
        col += 3

    ncols = 2 + len(TIMEFRAME_COLUMNS) * 3
    _style_header(ws, 1, ncols, _HEADER_FILL, _HEADER_FONT)
    _style_header(ws, 2, ncols, _SUB_FILL, _SUB_FONT)

    for r, item in enumerate(rows, start=3):
        ws.cell(row=r, column=1, value=item["code"]).alignment = _CENTER
        ws.cell(row=r, column=2, value=item["name"])
        c = 3
        for label, _s, _o in TIMEFRAME_COLUMNS:
            for key in ("K", "D", "J"):
                value = item.get(f"{label}_{key}")
                cell = ws.cell(row=r, column=c, value=value)
                cell.number_format = "0.00"
                c += 1

    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 14
    for c in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8


def write_flat_sheet(ws, headers: Sequence[str], rows: Sequence[Sequence],
                     widths: Optional[Sequence[int]] = None,
                     number_format: str = "0.0000",
                     col_formats: Optional[dict] = None) -> None:
    """需求 2/3/4 用的单行表头普通表。

    col_formats 按 1-based 列号覆盖默认格式 —— 同一张表里
    百分比列和金额列不能用同一个格式，否则「亿元」会被显示成百分数。
    """
    col_formats = col_formats or {}
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers), _HEADER_FILL, _HEADER_FONT)

    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if c > 2 and isinstance(value, (int, float)):
                cell.number_format = col_formats.get(c, number_format)

    ws.freeze_panes = "A2"
    if widths:
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
    else:
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = max(12, len(str(headers[c - 1])) + 4)


def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def save(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("已生成 %s", path)
    return path
